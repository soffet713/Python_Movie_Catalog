from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import date

today = date.today().strftime("%Y%m%d")
wb = Workbook()
wb.save(today + '_Movie_Catalog.xlsx')


def str_to_int_or_float(value):
    if isinstance(value, bool):
        return value
    try:
        return int(value)
    except ValueError:
        try:
            return float(value)
        except ValueError:
            return value


ws = wb.active
ws.title = 'Blu-Rays'
ws.append(['Title', 'Director', 'Genre', 'Release Year'])
movies = [{'title': '47 Ronin 3D', 'director': 'Carl Rinsch', 'genre': 'Action, Drama, Fantasy', 'releaseYear': '2013'},
          {'title': 'Alita: Battle Angel', 'director': 'Robert Rodriguez', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '2019'},
          {'title': 'Ant-man and the Wasp', 'director': 'Peyton Reed', 'genre': 'Action, Adventure, Comedy',
           'releaseYear': '2018'},
          {'title': 'The Avengers', 'director': 'Joss Whedon', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '2012'},
          {'title': 'Avengers: Infinity War', 'director': 'Anthony Russo, Joe Russo', 'genre':
              'Action, Adventure, Sci-Fi', 'releaseYear': '2018'},
          {'title': 'Avengers: Endgame', 'director': 'Anthony Russo, Joe Russo', 'genre': 'Action, Adventure, Drama',
           'releaseYear': '2019'},
          {'title': 'Baby Driver', 'director': 'Edgar Wright', 'genre': 'Action, Crime, Drama', 'releaseYear': '2017'},
          {'title': 'Batman v Superman: Dawn of Justice', 'director': 'Zack Snyder', 'genre':
              'Action, Adventure, Sci-Fi', 'releaseYear': '2016'},
          {'title': 'Black Panther', 'director': 'Ryan Coogler', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '2018'},
          {'title': 'Blade Runner 2049', 'director': 'Ryan Coogler', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '2018'},
          {'title': 'キャプテン・アメリカ/ザ・ファースト・アベンジャー', 'director': 'ジョー・ジョンストン', 'genre':
              'Action, Adventure, Sci-Fi', 'releaseYear': '2011'},
          {'title': 'Captain America: The Winter Soldier', 'director': 'Anthony Russo, Joe Russo', 'genre':
              'Action, Adventure, Sci-Fi', 'releaseYear': '2014'},
          {'title': 'Captain America: Civil War', 'director': 'Anthony Russo, Joe Russo', 'genre':
              'Action, Adventure, Sci-Fi', 'releaseYear': '2016'},
          {'title': 'Castle in the Sky', 'director': 'Hayao Miyazaki', 'genre': 'Animation, Adventure, Drama',
           'releaseYear': '1986'},
          {'title': 'Chappie', 'director': 'Neill Blomkamp', 'genre': 'Action, Crime, Drama',
           'releaseYear': '2015'},
          {'title': 'A Christmas Carol', 'director': 'Robert Zemeckis', 'genre': 'Animation, Drama, Family',
           'releaseYear': '2009'},
          {'title': 'Dawn of the Planet of the Apes', 'director': 'Matt Reeves', 'genre': 'Action, Adventure, Drama',
           'releaseYear': '2014'},
          {'title': 'Deadpool', 'director': 'Tim Miller', 'genre': 'Action, Adventure, Comedy',
           'releaseYear': '2016'},
          {'title': 'Deadpool 2', 'director': 'David Leitch', 'genre': 'Action, Adventure, Comedy',
           'releaseYear': '2018'},
          {'title': 'Detective Pikachu', 'director': 'Rob Letterman', 'genre': 'Adventure, Comedy, Family',
           'releaseYear': '2019'},
          {'title': 'Doctor Strange', 'director': 'Scott Derrickson', 'genre': 'Action, Adventure, Fantasy',
           'releaseYear': '2016'},
          {'title': 'Get Out', 'director': 'Jordan Peele', 'genre': 'Horror, Mystery, Thriller',
           'releaseYear': '2017'},
          {'title': 'Ex Machina', 'director': 'Alex Garland', 'genre': 'Drama, Mystery, Sci-Fi',
           'releaseYear': '2014'},
          {'title': 'The Fast and the Furious', 'director': 'Rob Cohen', 'genre': 'Action, Crime, Thriller',
           'releaseYear': '2001'},
          {'title': '2 Fast 2 Furious', 'director': 'John Singleton', 'genre': 'Action, Crime, Thriller',
           'releaseYear': '2003'},
          {'title': 'The Fast and the Furious: Tokyo Drift', 'director': 'Justin Lin', 'genre':
              'Action, Crime, Thriller', 'releaseYear': '2006'},
          {'title': 'Fast & Furious', 'director': 'Justin Lin', 'genre': 'Action, Crime, Thriller',
              'releaseYear': '2009'},
          {'title': 'Fast Five', 'director': 'Justin Lin', 'genre': 'Action, Adventure, Crime', 'releaseYear': '2011'},
          {'title': 'Fast & Furious 6', 'director': 'Justin Lin', 'genre': 'Action, Adventure, Crime',
           'releaseYear': '2013'},
          {'title': 'The Fate of the Furious', 'director': 'F. Gary Gray', 'genre': 'Action, Adventure, Crime',
           'releaseYear': '2017'},
          {'title': 'First Man', 'director': 'Damien Chazelle', 'genre': 'Biography, Drama, History',
           'releaseYear': '2018'},
          {'title': 'Furious 7 - Extended Edition', 'director': 'James Wan', 'genre': 'Action, Adventure, Crime',
           'releaseYear': '2015'},
          {'title': 'Godzilla', 'director': 'Gareth Edwards', 'genre': 'Action, Adventure, Sci-Fi', 'releaseYear':
              '2014'},
          {'title': 'Godzilla: King of the Monsters', 'director': 'Michael Dougherty', 'genre':
              'Action, Adventure, Fantasy', 'releaseYear': '2019'},
          {'title': 'Guardians of the Galaxy', 'director': 'James Gunn', 'genre': 'Action, Adventure, Comedy',
           'releaseYear': '2014'},
          {'title': 'Guardians of the Galaxy Vol. 2', 'director': 'James Gunn', 'genre': 'Action, Adventure, Comedy',
           'releaseYear': '2017'},
          {'title': 'Hobbs & Shaw', 'director': 'David Leitch', 'genre': 'Action, Adventure', 'releaseYear': '2019'},
          {'title': 'How the Universe Works', 'director': 'Mark Bridge', 'genre': 'Documentary', 'releaseYear': '2010'},
          {'title': 'ID4 - Independence Day', 'director': 'Roland Emmerich', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '1996'},
          {'title': 'Interstellar', 'director': 'Christopher Nolan', 'genre': 'Adventure, Drama, Sci-Fi',
           'releaseYear': '2014'},
          {'title': 'IT', 'director': 'Andy Muschietti', 'genre': 'Horror', 'releaseYear': '2017'},
          {'title': 'IT: Chapter Two', 'director': 'Andy Muschietti', 'genre': 'Drama, Fantasy, Horror',
           'releaseYear': '2019'},
          {'title': 'John Wick', 'director': 'Chad Stahelski, David Leitch (uncredited)', 'genre':
           'Action, Crime, Thriller', 'releaseYear': '2014'},
          {'title': 'John Wick: Chapter 2', 'director': 'Chad Stahelski', 'genre': 'Action, Crime, Thriller',
           'releaseYear': '2017'},
          {'title': 'John Wick: Chapter 3 - Parabellum', 'director': 'Chad Stahelski', 'genre':
              'Action, Crime, Thriller', 'releaseYear': '2019'},
          {'title': 'Joker', 'director': 'Todd Phillips', 'genre': 'Crime, Drama, Thriller',
           'releaseYear': '2019'},
          {'title': 'The Lion King', 'director': 'Roger Allers, Rob Minkoff', 'genre': 'Animation, Adventure, Drama',
           'releaseYear': '1994'},
          {'title': 'Logan', 'director': 'James Mangold', 'genre': 'Action, Drama, Sci-Fi', 'releaseYear': '2017'},
          {'title': 'Man of Steel', 'director': 'Zack Snyder', 'genre': 'Action, Adventure, Sci-Fi', 'releaseYear':
              '2013'},
          {'title': 'Marley', 'director': 'Kevin Macdonald', 'genre': 'Documentary, Biography, Music', 'releaseYear':
              '2012'},
          {'title': 'The Martian', 'director': 'Ridley Scott', 'genre': 'Adventure, Drama, Sci-Fi', 'releaseYear':
              '2015'},
          {'title': 'Mihimaru GT - Mihimalive3', 'director': 'Universal Music Japan', 'genre': 'Concert', 'releaseYear':
              '2010'},
          {'title': 'Mission Impossible: Fallout', 'director': 'Christopher McQuarrie', 'genre':
              'Action, Adventure, Thriller', 'releaseYear': '2018'},
          {'title': 'Nausicaa of the Valley of the Wind', 'director': 'Hayao Miyazaki', 'genre':
              'Animation, Adventure, Fantasy', 'releaseYear': '1984'},
          {'title': 'ナイトメアー・ビフォア・クリスマス', 'director': 'ヘンリー・セリック', 'genre':
              'Animation, Family, Fantasy', 'releaseYear': '1993'},
          {'title': 'Once Upon a Time in Hollywood', 'director': 'Quentin Tarantino', 'genre': 'Comedy, Drama',
           'releaseYear': '2019'},
          {'title': 'Pacific Rim', 'director': 'Guillermo del Toro', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '2013'},
          {'title': 'The Perfect Storm', 'director': 'Wolfgang Petersen', 'genre': 'Action, Adventure, Drama',
           'releaseYear': '2000'},
          {'title': 'Ponyo', 'director': 'Hayao Miyazaki', 'genre': 'Animation, Adventure, Comedy',
           'releaseYear': '2008'},
          {'title': 'Poseidon', 'director': 'Wolfgang Petersen', 'genre': 'Action, Adventure, Drama',
           'releaseYear': '2006'},
          {'title': 'Princess Mononoke', 'director': 'Hayao Miyazaki', 'genre': 'Animation, Adventure, Fantasy',
           'releaseYear': '1997'},
          {'title': 'Rise of the Planet of the Apes', 'director': 'Rupert Wyatt', 'genre': 'Action, Drama, Sci-Fi',
           'releaseYear': '2011'},
          {'title': 'Rocketman', 'director': 'Dexter Fletcher', 'genre': 'Biography, Drama, Music', 'releaseYear':
              '2019'},
          {'title': 'Rogue One: A Star Wars Story', 'director': 'Gareth Edwards', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '2016'},
          {'title': 'Scott Pilgrim vs. the World', 'director': 'Edgar Wright', 'genre': 'Action, Comedy, Fantasy',
           'releaseYear': '2010'},
          {'title': 'Shazam!', 'director': 'David F. Sandberg', 'genre': 'Action, Adventure, Comedy',
           'releaseYear': '2019'},
          {'title': 'Sicario', 'director': 'Denis Villeneuve', 'genre': 'Action, Crime, Drama', 'releaseYear': '2015'},
          {'title': 'Sleeping Beauty - Diamond Edition', 'director': 'Clyde Geronimi, Les Clark (uncredited)', 'genre':
              'Animation, Family, Fantasy', 'releaseYear': '1959'},
          {'title': 'Solo: A Star Wars Story', 'director': 'Ron Howard', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '2018'},
          {'title': 'Southpaw', 'director': 'Antoine Fuqua', 'genre': 'Drama, Sport', 'releaseYear': '2015'},
          {'title': 'Spiderman: Homecoming 3D', 'director': 'Jon Watts', 'genre': 'Action, Adventure, Sci-Fi',
              'releaseYear': '2017'},
          {'title': 'Spiderman: Far from Home', 'director': 'Jon Watts', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '2019'},
          {'title': 'Star Trek Beyond', 'director': 'Justin Lin', 'genre': 'Action, Adventure, Sci-Fi', 'releaseYear':
              '2016'},
          {'title': 'Star Trek Into Darkness 3D', 'director': 'J.J. Abrams', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '2013'},
          {'title': 'Star Wars - The Force Awakens', 'director': 'J.J. Abrams', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '2015'},
          {'title': 'Star Wars - The Last Jedi', 'director': 'Rian Johnson', 'genre': 'Action, Adventure, Fantasy',
           'releaseYear': '2017'},
          {'title': 'Star Wars - The Rise of Skywalker', 'director': 'J.J. Abrams', 'genre':
              'Action, Adventure, Fantasy', 'releaseYear': '2019'},
          {'title': 'Straight Outta Compton', 'director': 'F. Gary Gray', 'genre': 'Biography, Drama, History',
           'releaseYear': '2015'},
          {'title': 'Ted (Unrated)', 'director': 'Seth MacFarlane', 'genre': 'Comedy', 'releaseYear': '2012'},
          {'title': 'Teenage Mutant Ninja Turtles - The Original Movie', 'director': 'Steve Barron', 'genre':
              'Action, Adventure, Comedy', 'releaseYear': '1990'},
          {'title': 'Teenage Mutant Ninja Turtles II - The Secret of the Ooze', 'director': 'Michael Pressman', 'genre':
              'Action, Adventure, Comedy', 'releaseYear': '1991'},
          {'title': 'Teenage Mutant Ninja Turtles III - Turtles in Time', 'director': 'Stuart Gillard', 'genre':
              'Action, Adventure, Comedy', 'releaseYear': '1993'},
          {'title': 'Teenage Mutant Ninja Turtles 3D', 'director': 'Jonathan Liebesman', 'genre':
              'Action, Adventure, Comedy', 'releaseYear': '2014'},
          {'title': 'Thor: Ragnarok', 'director': 'Taika Waititi', 'genre': 'Action, Adventure, Comedy',
           'releaseYear': '2017'},
          {'title': 'Twister', 'director': 'Jan de Bont', 'genre': 'Action, Adventure, Thriller',
           'releaseYear': '1996'},
          {'title': 'War for the Planet of the Apes', 'director': 'Matt Reeves', 'genre':
              'Action, Adventure, Drama', 'releaseYear': '2017'},
          {'title': 'What We Do in the Shadows', 'director': 'Jemaine Clement, Taika Waititi', 'genre':
              'Comedy, Horror', 'releaseYear': '2014'},
          {'title': 'Whisper of the Heart', 'director': 'Yoshifumi Kondô', 'genre': 'Animation, Drama, Family',
           'releaseYear': '1995'},
          {'title': 'The Wolverine', 'director': 'James Mangold', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '2013'},
          {'title': 'Wonder Woman', 'director': 'Patty Jenkins', 'genre': 'Action, Adventure, Fantasy',
           'releaseYear': '2017'},
          {'title': 'X-Men: Days of Future Past', 'director': 'Bryan Singer', 'genre': 'Action, Adventure, Sci-Fi',
           'releaseYear': '2014'}
          ]


def adjust_columns(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column  # Get the column name you want to adjust
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width


def menu():
    user_input = input("Enter 'a' to add a movie, 'l' to see your movies, 'f' to find a movie, 's' to save movies to an"
                       " Excel file, and 'q' to quit: ")
    while user_input != 'q':
        if user_input == 'a':
            add_movie()
        elif user_input == 'l':
            show_movies()
        elif user_input == 'f':
            find_by = input("What property of the movie do you want to search by? ")
            looking_for = input(f"What {find_by} are you looking for? ")
            mymovie = find_movie(looking_for, lambda x: x[find_by])
            print(mymovie or 'No movies found.')
        elif user_input == 's':
            for movie in movies:
                ws.append(list(map(str_to_int_or_float, movie.values())))

            last_cell = ws.cell(row=ws.max_row, column=ws.max_column).coordinate
            movie_table = Table(displayName='MovieTable', ref='A1:{}'.format(last_cell))
            style = TableStyleInfo(name='TableStyleMedium6', showRowStripes=True)
            movie_table.tableStyleInfo = style
            ws.add_table(movie_table)
            adjust_columns(ws)
            wb.save(today + '_Movie_Catalog.xlsx')
        else:
            print('Unknown command-please try again.')
        user_input = input("\nEnter 'a' to add a movie, 'l' to see your movies, 'f' to find a movie, 's' to save movies"
                           " to an Excel file, and 'q' to quit: ")


def add_movie():
    title = input("Enter the movie title: ")
    director = input("Enter the movie director: ")
    genre = input("Enter the genre of the movie: ")
    year = input("Enter the movie release year: ")

    movies.append({
        'title': title,
        'director': director,
        'genre': genre,
        'releaseYear': year
    })


def show_movies():
    for movie in movies:
        show_movie_details(movie)


def show_movie_details(movie):
    print(f"Title: {movie['title']}")
    print(f"Director: {movie['director']}")
    print(f"Genre: {movie['genre']}")
    print(f"Release year: {movie['releaseYear']}")


def find_movie(expected, finder):
    found = []
    for xmovie in movies:
        if finder(xmovie) == expected:
            found.append(xmovie)
    return found


menu()
