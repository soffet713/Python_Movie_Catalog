from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from Movie_Catalog import str_to_int_or_float, adjust_columns
from datetime import date

today = date.today().strftime("%Y%m%d")
# set file path
file_path = 'C:\\Users\\SeanMac\\PycharmProjects\\Python_Test01\\20200409_Movie_Catalog.xlsx'
wb = load_workbook(file_path)
wb.create_sheet('Blu-Ray Box Sets')

wb.active = wb['Blu-Ray Box Sets']
ws = wb.active

ws.append(['Title', 'Director', 'Genre', 'Release Year'])
box_sets = [{'title': 'The Dark Knight Trilogy', 'director': 'Christopher Nolan', 'genre': 'Action, Adventure, Crime',
             'releaseYear': '2012'},
            {'title': 'Game of Thrones - The Complete First Season', 'director': 'David Benioff, D.B. Weiss', 'genre':
                'Action, Adventure, Drama', 'releaseYear': '2011'},
            {'title': 'Game of Thrones - The Complete Second Season', 'director': 'David Benioff, D.B. Weiss', 'genre':
                'Action, Adventure, Drama', 'releaseYear': '2012'},
            {'title': 'Game of Thrones - The Complete Third Season', 'director': 'David Benioff, D.B. Weiss', 'genre':
                'Action, Adventure, Drama', 'releaseYear': '2013'},
            {'title': 'Game of Thrones - The Complete Fourth Season', 'director': 'David Benioff, D.B. Weiss', 'genre':
                'Action, Adventure, Drama', 'releaseYear': '2014'},
            {'title': 'Game of Thrones - The Complete Fifth Season', 'director': 'David Benioff, D.B. Weiss', 'genre':
                'Action, Adventure, Drama', 'releaseYear': '2015'},
            {'title': 'Game of Thrones - The Complete Sixth Season', 'director': 'David Benioff, D.B. Weiss', 'genre':
                'Action, Adventure, Drama', 'releaseYear': '2016'},
            {'title': 'Game of Thrones - The Complete Seventh Season', 'director': 'David Benioff, D.B. Weiss', 'genre':
                'Action, Adventure, Drama', 'releaseYear': '2017'},
            {'title': 'Game of Thrones - The Complete Eighth Season', 'director': 'David Benioff, D.B. Weiss', 'genre':
                'Action, Adventure, Drama', 'releaseYear': '2019'},
            {'title': 'Harry Potter - Complete Film Collection', 'director':
                'Chris Columbus, Alfonso Cuarón, Mike Newell, David Yates', 'genre': 'Action, Adventure, Family',
             'releaseYear': '2011'},
            {'title': 'The Lord of the Rings - The Motion Picture Trilogy', 'director': 'Peter Jackson', 'genre':
                'Adventure, Drama, Fantasy', 'releaseYear': '2011'},
            {'title': 'Star Wars - The Complete Saga', 'director': 'George Lucas, Irvin Kershner, Richard Marquand',
             'genre': 'Action, Adventure, Fantasy', 'releaseYear': '2011'}
            ]


def add_box_sets(set_list):
    for boxset in set_list:
        ws.append(list(map(str_to_int_or_float, boxset.values())))

    last_cell = ws.cell(row=ws.max_row, column=ws.max_column).coordinate
    box_set_table = Table(displayName='BoxSetTable', ref='A1:{}'.format(last_cell))
    style = TableStyleInfo(name='TableStyleMedium5', showRowStripes=True)
    box_set_table.tableStyleInfo = style
    ws.add_table(box_set_table)
    adjust_columns(ws)
    wb.save(today + '_Movie_Catalog.xlsx')


add_box_sets(box_sets)

wb.create_sheet('DVDs & Box Sets')
wb.active = wb['DVDs & Box Sets']
ws = wb.active

ws.append(['Title', 'Director', 'Genre', 'Release Year'])
dvds = [{'title': 'Akira', 'director': 'Katsuhiro Ôtomo', 'genre': 'Animation, Drama, Sci-Fi', 'releaseYear': '1988'},
          {'title': 'Appleseed', 'director': 'Shinji Aramaki, Steven Foster', 'genre': 'Animation, Action, Adventure',
           'releaseYear': '2004'},
          {'title': 'The Art of Rap', 'director': 'Ice-T, Andy Baybutt', 'genre': 'Documentary, Music',
           'releaseYear': '2012'},
          {'title': 'Blood: The Last Vampire', 'director': 'Hiroyuki Kitakubo', 'genre': 'Animation, Action, Horror',
           'releaseYear': '2000'},
          {'title': 'The Boondock Saints', 'director': 'Troy Duffy', 'genre': 'Action, Crime, Thriller',
           'releaseYear': '1999'},
          {'title': 'Case Closed: The Time Bombed Skyscraper', 'director': 'Kenji Kodama', 'genre':
              'Animation, Adventure, Crime', 'releaseYear': '1997'},
          {'title': 'CB4 - The Movie', 'director': 'Tamra Davis', 'genre': 'Comedy, Music', 'releaseYear': '1993'},
          {'title': 'Dave Chappelle - Killin\' Them Softly', 'director': 'Stan Lathan', 'genre': 'Comedy Stand-up',
           'releaseYear': '2000'},
          {'title': 'Collateral', 'director': 'Michael Mann', 'genre': 'Crime, Drama, Thriller',
           'releaseYear': '2004'},
          {'title': 'Don\'t Be A Menace (Unrated)', 'director': 'Paris Barclay', 'genre': 'Comedy, Crime',
           'releaseYear': '1996'},
          {'title': 'Donnie Darko', 'director': 'Richard Kelly', 'genre': 'Drama, Mystery, Sci-Fi',
           'releaseYear': '2001'},
          {'title': 'Dragonball Evolution', 'director': 'James Wong', 'genre': 'Action, Adventure, Family',
           'releaseYear': '2009'},
          {'title': 'Dragonball Z - The History of Trunks / Bardock the Father of Goku', 'director':
              'Yoshihiro Ueda, Daisuke Nishio, Mitsuo Hashimoto', 'genre': 'Animation, Drama, Fantasy',
           'releaseYear': '1993'},
          {'title': 'Drumline', 'director': 'Charles Stone III', 'genre': 'Comedy, Drama, Romance',
           'releaseYear': '2002'},
          {'title': 'Family Guy Presents - Stewie Griffin: The Untold Story', 'director': 'Pete Michels, Peter Shin',
           'genre': 'Animation, Adventure, Comedy', 'releaseYear': '2005'},
          {'title': 'Grandma\'s Boy (Unrated)', 'director': 'Nicholaus Goossen', 'genre': 'Comedy',
           'releaseYear': '2006'},
          {'title': 'Grave of the Fireflies', 'director': 'Isao Takahata', 'genre': 'Animation, Drama, War',
           'releaseYear': '1988'},
          {'title': 'Head of State', 'director': 'Chris Rock', 'genre': 'Comedy', 'releaseYear': '2003'},
          {'title': 'How High', 'director': 'Jesse Dylan', 'genre': 'Comedy, Fantasy', 'releaseYear': '2001'},
          {'title': 'Howl\'s Moving Castle', 'director': 'Hayao Miyazaki', 'genre': 'Animation, Adventure, Family',
           'releaseYear': '2004'},
          {'title': 'Hustle & Flow', 'director': 'Craig Brewer', 'genre': 'Crime, Drama, Music', 'releaseYear': '2005'},
          {'title': 'I Tried', 'director': 'Rich Newey', 'genre': 'Drama, Music', 'releaseYear': '2007'},
          {'title': 'Idlewild', 'director': 'Bryan Barber', 'genre': 'Crime, Drama, Musical',
           'releaseYear': '2006'},
          {'title': 'Irish Jam', 'director': 'John Eyres', 'genre': 'Comedy', 'releaseYear': '2006'},
          {'title': 'Jacked Up', 'director': 'Timothy Wayne Folsome', 'genre': 'Crime, Drama, Thriller',
           'releaseYear': '2001'},
          {'title': 'The Last Samurai', 'director': 'Edward Zwick', 'genre': 'Action, Drama, War',
           'releaseYear': '2003'},
          {'title': 'Letters from Iwo Jima', 'director': 'Clint Eastwood', 'genre': 'Action, Adventure, Drama',
              'releaseYear': '2006'},
          {'title': 'Liar Liar', 'director': 'Tom Shadyac', 'genre': 'Comedy, Fantasy', 'releaseYear': '1997'},
          {'title': 'The Longest Yard', 'director': 'Peter Segal', 'genre': 'Comedy, Crime, Sport',
           'releaseYear': '2005'},
          {'title': 'Malibu\'s Most Wanted', 'director': 'John Whitesell', 'genre': 'Comedy, Crime',
           'releaseYear': '2003'},
          {'title': 'Man on Fire', 'director': 'Tony Scott', 'genre': 'Action, Crime, Drama',
           'releaseYear': '2004'},
          {'title': 'Metropolis (メトロポリス)', 'director': 'Rintaro', 'genre': 'Animation, Adventure, Drama',
           'releaseYear': '2001'},
          {'title': 'My Neighbor Totoro', 'director': 'Hayao Miyazaki', 'genre': 'Animation, Family, Fantasy',
           'releaseYear': '1988'},
          {'title': 'Nausicaa of the Valley of the Wind', 'director': 'Hayao Miyazaki', 'genre':
              'Animation, Adventure, Fantasy', 'releaseYear': '1984'},
          {'title': 'Ninja Scroll', 'director': 'Yoshiaki Kawajiri', 'genre': 'Animation, Action, Adventure',
           'releaseYear': '1993'},
          {'title': 'Porco Rosso', 'director': 'Hayao Miyazaki', 'genre': 'Animation, Adventure, Comedy',
           'releaseYear': '1992'},
          {'title': 'The Princess Blade', 'director': 'Shinsuke Sato', 'genre': 'Action, Sci-Fi',
           'releaseYear': '2001'},
          {'title': 'Princess Mononoke', 'director': 'Hayao Miyazaki', 'genre': 'Animation, Adventure, Fantasy',
           'releaseYear': '1997'},
          {'title': 'R.O.D. - Read Or Die', 'director': 'Kôji Masunari', 'genre': 'Animation, Short, Adventure',
           'releaseYear': '2001'},
          {'title': 'Seven Samurai', 'director': 'Akira Kurosawa', 'genre': 'Action, Adventure, Drama',
           'releaseYear': '1954'},
          {'title': 'Spirited Away', 'director': 'Hayao Miyazaki', 'genre': 'Animation, Adventure, Family',
           'releaseYear': '2001'},
          {'title': 'Steamboy', 'director': 'Katsuhiro Ôtomo', 'genre': 'Animation, Action, Adventure',
           'releaseYear': '2004'},
          {'title': 'Tokyo Babylon', 'director': 'Kôichi Chigira', 'genre': 'Animation, Drama, Fantasy',
           'releaseYear': '1992'},
          {'title': 'Versus', 'director': 'Ryûhei Kitamura', 'genre': 'Action, Comedy, Drama', 'releaseYear': '2000'},
          {'title': 'War Room', 'director': 'Alex Kendrick', 'genre': 'Drama', 'releaseYear': '2015'},
          {'title': 'Whisper of the Heart', 'director': 'Yoshifumi Kondô', 'genre': 'Animation, Drama, Family',
           'releaseYear': '1995'},
          {'title': 'Zatoichi: The Blind Swordsman', 'director': 'Takeshi Kitano', 'genre': 'Action, Comedy, Crime',
           'releaseYear': '2003'},
          {'title': 'ALWAYS三丁目の夕日/続・三丁目の夕日', 'director': '山崎貴', 'genre': 'Drama, Family',
           'releaseYear': '2008'},
          {'title': 'いま、会いにゆきます', 'director': 'Nobuhiro Doi', 'genre': 'Drama, Fantasy, Romance',
           'releaseYear': '2004'},
          {'title': 'GO', 'director': 'Isao Yukisada', 'genre': 'Drama', 'releaseYear': '2001'},
          {'title': 'NANA', 'director': 'Kentarô Ohtani', 'genre': 'Drama, Music, Romance', 'releaseYear': '2015'},
          {'title': 'ナイトメアー・ビフォア・クリスマス', 'director': 'Henry Selick', 'genre': 'Animation, Family, Fantasy',
           'releaseYear': '1993'},
          {'title': 'ダウンタウンのガキの使いやあらへんで！！ 17罰 23 上', 'director': 'ダウンタウン, 山崎邦正, ココリコ',
           'genre': 'Television, Comedy', 'releaseYear': '2011'},
          {'title': 'ダウンタウンのガキの使いやあらへんで！！ 17罰 23 下', 'director': 'ダウンタウン, 山崎邦正, ココリコ',
           'genre': 'Television, Comedy', 'releaseYear': '2011'},
          {'title': 'ダウンタウンのガキの使いやあらへんで！！ 18罰 30 上', 'director': 'ダウンタウン, 山崎邦正, ココリコ',
           'genre': 'Television, Comedy', 'releaseYear': '2012'},
          {'title': 'ダウンタウンのガキの使いやあらへんで！！ 18罰 30 下', 'director': 'ダウンタウン, 山崎邦正, ココリコ',
           'genre': 'Television, Comedy', 'releaseYear': '2012'},
          {'title': 'Orange Range - Live Tour 006 \'FANTAZICAL\'', 'director': 'ソニーレコード', 'genre':
              'J-Pop, Concert', 'releaseYear': '2007'},
          {'title': '倖田來未 - Secret First Class Limited Live', 'director': 'APEX Marketing', 'genre':
              'J-Pop, Concert', 'releaseYear': '2005'},
          {'title': 'コブクロ - Live Tour \'06 "Way Back to Tomorrow" FINAL', 'director': 'Warner Music Japan', 'genre':
              'J-Pop, Concert', 'releaseYear': '2007'},
          {'title': 'SOFFet - Turn the Jam\'s Key Tour', 'director': 'SPACE SHOWER MUSIC', 'genre': 'J-Pop, Concert',
           'releaseYear': '2012'},
          {'title': '', 'director': '', 'genre': '', 'releaseYear': ''},
          {'title': '', 'director': '', 'genre': '', 'releaseYear': ''},
          {'title': 'DVD Box Sets', 'director': '', 'genre': '', 'releaseYear': ''},
          {'title': 'Chappelle\'s Show - The Complete Series', 'director': 'Neal Brennan, Dave Chappelle',
           'genre': 'Comedy, Music, TV', 'releaseYear': '2012'},
          {'title': 'Family Guy - Volume One', 'director': 'Seth MacFarlane, David Zuckerman', 'genre':
              'Animation, Comedy, TV', 'releaseYear': '2010'},
          {'title': 'Family Guy - Volume Three', 'director': 'Seth MacFarlane, David Zuckerman', 'genre':
              'Animation, Comedy, TV', 'releaseYear': '2010'},
          {'title': 'Psych - The Complete Series', 'director': 'Steve Franks', 'genre': 'Comedy, Crime, Mystery',
           'releaseYear': '2014'}
        ]


def add_dvds(dvd_list):
    for dvd in dvd_list:
        ws.append(list(map(str_to_int_or_float, dvd.values())))

    last_cell = ws.cell(row=ws.max_row, column=ws.max_column).coordinate
    dvd_set_table = Table(displayName='DVDSetTable', ref='A1:{}'.format(last_cell))
    style = TableStyleInfo(name='TableStyleMedium7', showRowStripes=True)
    dvd_set_table.tableStyleInfo = style
    ws.add_table(dvd_set_table)
    adjust_columns(ws)
    wb.save(today + '_Movie_Catalog.xlsx')


add_dvds(dvds)
